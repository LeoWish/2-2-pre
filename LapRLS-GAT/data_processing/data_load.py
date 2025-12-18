import openpyxl
import numpy as np
import pandas as pd


np.set_printoptions(threshold=np.inf)


def load_data(path):

    data_wb = openpyxl.load_workbook(path)
    sheets = {
        'lb_KEGG': data_wb['KEGG_lb'],
        'lb_GPA': data_wb['GPA_lb'],
        'lb_IS': data_wb['IS_lb'],
        'st_KEGG': data_wb['KEGG_st'],
        'st_GPA': data_wb['GPA_st'],
        'st_IS': data_wb['IS_st'],
        'A': data_wb['A'],
    }

    data_pairs = []
    interaction_sheet = sheets['A']
    for row in range(2, interaction_sheet.max_row + 1):
        compound = interaction_sheet.cell(row, 1).value
        for col in range(2, interaction_sheet.max_column + 1):
            data = interaction_sheet.cell(1, col).value
            data_pairs.append(f"{compound}&{data}")

    def load_sheet_values(sheet):
        values = []
        for row in range(2, sheet.max_row + 1):  # 从第二行开始读取
            row_values = [sheet.cell(row, col).value for col in range(2, sheet.max_column + 1)]  # 从第二列开始读取
            values.append(row_values)
        return np.array(values)

    # 加载各工作表数据
    data_lb_KEGG_values = load_sheet_values(sheets['lb_KEGG'])
    data_lb_GPA_values = load_sheet_values(sheets['lb_GPA'])
    data_lb_IS_values = load_sheet_values(sheets['lb_IS'])
    data_st_KEGG_values = load_sheet_values(sheets['st_KEGG'])
    data_st_GPA_values = load_sheet_values(sheets['st_GPA'])
    data_st_IS_values = load_sheet_values(sheets['st_IS'])
    data_interaction_values = load_sheet_values(sheets['A'])


    return (
        data_lb_KEGG_values,
        data_lb_GPA_values,
        data_lb_IS_values,
        data_st_KEGG_values,
        data_st_GPA_values,
        data_st_IS_values,
        data_interaction_values,
        data_pairs
    )

